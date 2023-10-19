
#Användbara Paket
#För att hämta datan
import requests
#Tider och datum
import datetime
#För att kunna kolla om filen existens
import os
#Använd för att formatera och överföra till excel fil, as op för att jag lat
import openpyxl as op
"""
Anteckningar
Behöver i kolumnerna i excel:
När den hämtades: Datetime.now
longitud: Finns under Geometry, coordinates
Latitud: Finns under Geometry, coordinates
Datum = Vilket datum, Ta ur från validTime
Hour: vilken timme datan gäller. 
Temperatur: Float, t = temperatur
RainOrSnow: bool, pcat == 0 leder till False, pcat == 1-6 -> True
provider: Alltid SMHI
"""
#Skapar excelfilen om den inte finns
def excel_creation():
    #Om den inte finns, Behöver ingen else. Skipas då den redan har filen
    if not os.path.isfile('Väder_data.xlsx'):
        #Skapa "Boken"
        workbook = op.Workbook()
        #Skapa sidan för datan
        worksheet = workbook.active
        #Sätt titel på sidan
        worksheet.title = 'Väder Data'
        #Sätt titlar för all hämtad data
        worksheet['A1'] = "Skapad"
        worksheet['B1'] = "Longitud"
        worksheet['C1'] = "Latitud"
        worksheet['D1'] = "Datum"
        worksheet['E1'] = "Timma"
        worksheet['F1'] = "Temperatur"
        worksheet['G1'] = "Nederbörd"
        worksheet['H1'] = "Källa"
        #Spara så att filen faktiskt sparas
        workbook.save ('Väder_data.xlsx')


#Skapande av Huvud menyn och input av alternativ
def main ():
    while True:
        print ("1. Hämta senaste datan från SMHI")
        print ("2. Skriv ut prognosen")
        print ("9. Avsluta programmet")
        val = input (" > ")
        #Leder till insamling av data
        if val == "1":
            #Url för SMHI data
            url = "https://opendata-download-metfcst.smhi.se/api/category/pmp3g/version/2/geotype/point/lon/18.0215/lat/59.3099/data.json"
            
            #Hämta datan från SMHI
            SMHI_hämta_data = requests.get (url)
            
            #Parse datan ifrån JSON
            SMHI_data = SMHI_hämta_data.json()
            
            #Ladda xlsx filen med sidan som datan ska ligga på
            workbook = op.load_workbook ('Väder_data.xlsx')
            sheet = workbook.active
            nutid = datetime.datetime.now()
            #Repetera 24 gånger för alla värden
            for repetition in range (1,25):
                
                #Hitta sista raden, +1 så att den inte överskriver
                sista_rad = sheet.max_row + 1 
                
                #Skapad
                sheet.cell (row=sista_rad, column=1, value=nutid)

                #Longitud
                longitud = SMHI_data ['geometry']['coordinates'][0][0]
                sheet.cell (row=sista_rad, column=2, value=longitud)

                #Latitud
                latitud = SMHI_data ['geometry']['coordinates'][0][1]
                sheet.cell (row=sista_rad, column=3, value=latitud)

                #validTime i sträng format
                validTime_str = SMHI_data['timeSeries'][repetition]['validTime']
                validTime_dt = datetime.datetime.strptime(validTime_str, "%Y-%m-%dT%H:%M:%SZ") + datetime.timedelta(hours=2)
                
                #Datum    
                datum = validTime_dt.date()
                sheet.cell (row=sista_rad, column=4, value=datum)
                
                #Timma
                timma = validTime_dt.hour
                sheet.cell (row=sista_rad, column=5, value=timma)
                
                #Nå in i parameterarna
                for parameters in SMHI_data['timeSeries'][repetition]['parameters']:
                    #Hitta temperatur
                    if parameters['name'] == 't':
                        #Få ur värdet ur listan
                        temperatur = parameters['values'][0]
                        sheet.cell (row=sista_rad, column=6, value=temperatur)
                    
                    #Hitta nederbörd
                    if parameters['name'] == 'pcat':
                        nederbörd = parameters ['values']
                        #Ändra till Bool
                        if nederbörd[0] == 0:
                            sheet.cell (row=sista_rad, column=7, value=False)
                        elif nederbörd[0] >= 1:
                            sheet.cell (row=sista_rad, column=7, value=True)

            
                #Provider
                sheet.cell (row=sista_rad, column=8, value="SMHI")

            #Spara den ändrade datan
            workbook.save ('Väder_data.xlsx')
            print ("Data hämtad och sparad till Väder_data.xlsx")
            


        #Printa ut från Excel fil
        elif val == "2":
            #ladda in workbook och sheet
            workbook = op.load_workbook ('Väder_data.xlsx')
            sheet = workbook.active
            #Skapa stop och start värde för läsningsramen
            läs_stop = sheet.max_row
            läs_start = läs_stop - 23
            
            #Printa gällande datum
            print_datum = sheet.cell (row=läs_start, column=4).value
            print_datum = datetime.datetime.strftime (print_datum,"%Y-%m-%d")
            print (f"Prognos from SMHI {print_datum}:")
            #Läs raderna, +1 för stopvärde
            for rad in range (läs_start, läs_stop+1):
                #De olika Värdena som ska inkluderas
                print_timma = sheet.cell (row=rad, column=5).value
                print_temperatur = sheet.cell (row=rad, column=6).value

                #Fixa nederbörd till rätt visat i print
                if sheet.cell (row=rad, column=7).value == True:
                    print_nederbörd = "Nederbörd"
                else:
                    print_nederbörd = "Ingen nederbörd"
                
                #Print för terminal
                print (f"{print_timma:02}:00 {print_temperatur} Grader {print_nederbörd}")
                

        
        #Avsluta Programmet
        elif val == "9":
            break
        
        #Stoppande av program avslut vid fel input
        else:
            print ("Ej igenkänd input. Försök igen.")






if __name__ == "__main__":
    #Kolla om xlsx fil finns eller inte
    excel_creation()
    
    #Självaste programmet
    main()


